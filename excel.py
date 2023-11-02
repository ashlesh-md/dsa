import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage
from openpyxl.styles import Font, Alignment
import json

# Load data from the JSON file
with open('api_data.json', 'r') as json_file:
    api_data = json.load(json_file)

# Create a DataFrame from the loaded data
df = pd.DataFrame(api_data)

# Create a new Excel workbook
wb = Workbook()
ws = wb.active

# Set the width for cells A, B, C, and D
for col in "ABCD":
    ws.column_dimensions[col].width = 50

# Define the font style with size 14
font_style = Font(size=14)

# Write the DataFrame to the Excel sheet
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
    for c_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        # Apply the defined font style
        cell.font = font_style
        if r_idx == 1:
            cell.alignment = cell.alignment.copy(vertical="top")
        elif r_idx > 1 and c_idx == 4:
            image_filename = os.path.join("images", value)
            img = PILImage.open(image_filename)
            img_path = f'temp_image_{r_idx}.png'
            img.save(img_path)
            img_obj = ExcelImage(img_path)
            img_obj.width = 400
            img_obj.height = 720
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.alignment = Alignment(horizontal="center", vertical="center")  # Center image
            ws.add_image(img_obj, cell.coordinate)
        else:
            cell.alignment = cell.alignment.copy(wrapText=True, vertical="top")

# Set the row height for all rows except the first row
for r_idx in range(2, ws.max_row + 1):
    ws.row_dimensions[r_idx].height = 550

# Save the workbook to a file
wb.save('api_data_with_images.xlsx')
